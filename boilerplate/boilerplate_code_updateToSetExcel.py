from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import time
import traceback
import getpass
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 웹드라이버 설정
driver = webdriver.Chrome()

# 로그인 페이지로 이동
driver.get('https://gw.com2us.com/')

# 로그인 처리
username_input = driver.find_element(By.ID, 'Username')
password_input = driver.find_element(By.ID, 'Password')

username = input('아이디를 입력하세요: ')
password = getpass.getpass('비밀번호를 입력하세요: ')

username_input.send_keys(username)
password_input.send_keys(password)

login_button = driver.find_element(By.CLASS_NAME, 'btnLogin')
login_button.click()

time.sleep(5)

# 결재 페이지로 이동
driver.get('https://gw.com2us.com/emate_appro/appro_complete_2024_link.nsf/wfmViaView?readform&viewname=view052&vctype=a&viewcode=2024082903')

# 페이지 로딩 대기
WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//tr[contains(@class, "dhx_skyblue")]')))

# 페이지 소스 저장 (디버깅용)
with open('page_source.html', 'w', encoding='utf-8') as f:
    f.write(driver.page_source)

# 게시글 목록에서 데이터 추출
posts = driver.find_elements(By.XPATH, '//tr[@class="odd_dhx_skyblue " or @class="ev_dhx_skyblue "]')
print(f"게시글 수: {len(posts)}")

data_list = []

for i in range(len(posts)):
    # 게시글 목록을 다시 가져옵니다.
    posts = driver.find_elements(By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]')
    post = posts[i]

    try:
        # 해당 행의 모든 td 요소를 가져옵니다.
        tds = post.find_elements(By.TAG_NAME, 'td')
        
        # 결재일 추출 (결재일이 있는 td 요소 선택)
        결재일_td = tds[6]  # 결재일이 7번째 컬럼에 위치
        결재일_text = 결재일_td.text.strip()

        # 년, 월, 일 추출
        년 = 결재일_text.split('-')[0]
        월 = str(int(결재일_text.split('-')[1]))
        일 = str(int(결재일_text.split('-')[2]))

        print(f"결재일: {결재일_text}, 년: {년}, 월: {월}, 일: {일}")

        # 신청자 추출 (신청자가 있는 td 요소 선택)
        신청자_td = tds[4]  # 신청자가 5번째 컬럼에 위치
        신청자_span = 신청자_td.find_element(By.TAG_NAME, 'span')
        신청자 = 신청자_span.text.strip()

        print(f"신청자: {신청자}")

    except Exception as e:
        print(f"목록에서 데이터 추출 중 오류 발생: {e}")
        결재일_text = 년 = 월 = 일 = 신청자 = ''

    # 요소가 화면에 보이도록 스크롤합니다.
    driver.execute_script("arguments[0].scrollIntoView();", post)

    # 클릭 가능할 때까지 대기합니다.
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]')))

    # 게시글 클릭하여 팝업 열기
    post.click()

    # 새로운 창으로 전환
    WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
    driver.switch_to.window(driver.window_handles[-1])

    # 필요한 시간 대기 (페이지 로딩)
    time.sleep(3)

    try:
        # 현재 창 제목 출력
        print(f"현재 창 제목: {driver.title}")

        # 법인명 추출
        법인명_element = driver.find_elements(By.ID, 'titleLabel')
        if 법인명_element:
            법인명 = 법인명_element[0].text.strip()
        else:
            print("법인명 요소를 찾을 수 없습니다.")
            법인명 = ''

        # 문서번호 추출
        문서번호_element = driver.find_elements(By.XPATH, '//th[text()="문서번호"]/following-sibling::td[1]')
        if 문서번호_element:
            문서번호 = 문서번호_element[0].text.strip()
        else:
            print("문서번호 요소를 찾을 수 없습니다.")
            문서번호 = ''

        # 제목 추출
        제목_element = driver.find_elements(By.CSS_SELECTOR, 'td.approval_text')
        if 제목_element:
            제목_text = 제목_element[0].text.strip()
            제목 = 제목_text.replace(법인명, '').strip()
        else:
            print("제목 요소를 찾을 수 없습니다.")
            제목 = ''

        # 합의 담당자 추출
        합의담당자_element = driver.find_elements(By.XPATH, '//tr[@class="name"]/td[@class="td_point"]')
        if 합의담당자_element:
            합의담당자 = 합의담당자_element[0].text.strip()
        else:
            print("합의 담당자 요소를 찾을 수 없습니다.")
            합의담당자 = ''

        # 링크 추출
        링크 = driver.current_url

        # 추출된 데이터 출력
        print(f"법인명: {법인명}")
        print(f"문서번호: {문서번호}")
        print(f"제목: {제목}")
        print(f"합의 담당자: {합의담당자}")
        print(f"링크: {링크}")

        # 데이터 저장
        data = {
        '결재일': 결재일_text,
        '년': 년,
        '월': 월,
        '일': 일,
        '주차': '',          # 빈 문자열 할당
        '법인명': 법인명,
        '문서번호': 문서번호,
        '제목': 제목,
        '업무 유형': '',      # 빈 문자열 할당
        '추출 위치': '',      # 빈 문자열 할당
        '담당 부서': '',      # 빈 문자열 할당
        '신청자': 신청자,
        '합의 담당자': 합의담당자,
        '링크': 링크,
        '진행 구분': ''       # 빈 문자열 할당
    }
        data_list.append(data)
    except Exception as e:
        print(f"데이터 추출 중 오류 발생: {e}")
        traceback.print_exc()

    # 창 닫기 및 원래 창으로 전환
    driver.close()
    driver.switch_to.window(driver.window_handles[0])

    # 잠시 대기
    time.sleep(2)

# 데이터프레임 생성
df = pd.DataFrame(data_list)


######################################3안 시작##############################################

# 기존 엑셀 파일 불러오기
excel_file = r'C:\Users\PHJ\output\개인정보 운영대장.xlsx'
wb = load_workbook(excel_file)
ws = wb['개인정보 추출 및 이용 관리']  # 데이터가 입력될 워크시트 이름

# 'NO' 열에서 데이터가 있는 마지막 행 찾기
last_row = ws.max_row
while last_row >= 6:  # 데이터가 시작되는 행 번호는 6
    if ws.cell(row=last_row, column=2).value is not None:
        break
    last_row -= 1

# 새로운 데이터 입력 시작 행
if last_row < 6:
    start_row = 6  # 데이터 시작 행
else:
    start_row = last_row + 1

# 다음 'NO' 값 결정
if last_row >= 6:
    last_no = ws.cell(row=last_row, column=2).value
    if isinstance(last_no, int):
        next_no = last_no + 1
    else:
        next_no = 1
else:
    next_no = 1

# 데이터프레임의 열 순서 조정 (엑셀의 열 순서와 일치하도록)
df = df[['결재일', '년', '월', '일', '주차', '법인명', '문서번호', '제목', '업무 유형', '추출 위치', '담당 부서', '신청자', '합의 담당자', '링크', '진행 구분']]

# 열 매핑 설정 (데이터프레임 열 이름과 엑셀 열 인덱스 매핑)
column_mapping = {
    '결재일': 3,     # C
    '년': 4,         # D
    '월': 5,         # E
    '일': 6,         # F
    '주차': 7,       # G
    '법인명': 8,     # H
    '문서번호': 9,   # I
    '제목': 10,      # J
    '업무 유형': 11, # K
    '추출 위치': 12, # L
    '담당 부서': 13, # M
    '신청자': 14,    # N
    '합의 담당자': 15, # O
    '링크': 16,      # P
    '진행 구분': 17   # Q
}

# 데이터프레임을 엑셀 워크시트에 쓰기
for idx, row in df.iterrows():
    # 'NO' 값 입력 (열 인덱스 2는 B열)
    ws.cell(row=start_row, column=2, value=next_no)
    next_no += 1

    # 각 열에 데이터 입력
    for col_name, col_idx in column_mapping.items():
        value = row[col_name]
        ws.cell(row=start_row, column=col_idx, value=value)
    start_row += 1

# 엑셀 파일 저장
wb.save(excel_file)

######################################3안 끝##############################################


######################################1안 시작##############################################
# # 기존 엑셀 파일 불러오기
# excel_file = r'C:\Users\PHJ\output\개인정보 운영대장.xlsx'
# wb = load_workbook(excel_file)
# ws = wb['개인정보 추출 및 이용 관리']  # 데이터가 입력될 워크시트 이름

# # 'NO' 열에서 데이터가 있는 마지막 행 찾기
# last_row = ws.max_row
# while last_row > 1:
#     if ws.cell(row=last_row, column=1).value is not None:
#         break
#     last_row -= 1

# # 새로운 데이터 입력 시작 행
# start_row = last_row + 1

# # 다음 'NO' 값 결정
# if last_row > 1:
#     last_no = ws.cell(row=last_row, column=1).value
#     if isinstance(last_no, int):
#         next_no = last_no + 1
#     else:
#         next_no = 1
# else:
#     next_no = 1

# # 데이터프레임의 열 순서 조정 (엑셀의 열 순서와 일치하도록)
# df = df[['결재일', '년', '월', '일', '주차', '법인명', '문서번호', '제목', '업무 유형', '추출 위치', '담당 부서', '신청자', '합의 담당자', '링크', '진행 구분']]

# # 데이터프레임을 엑셀 워크시트에 쓰기
# for idx, row in df.iterrows():
#     # 'NO' 값 입력
#     ws.cell(row=start_row, column=1, value=next_no)
#     next_no += 1

#     # 데이터 입력 (열은 2부터 시작)
#     for col_idx, (col_name, value) in enumerate(row.items(), start=6):
#         ws.cell(row=start_row, column=col_idx, value=value)
#     start_row += 1

# # 엑셀 파일 저장
# wb.save(excel_file)

######################################1안 끝##############################################


######################################2안 시작##############################################

# # 기존 엑셀 파일 불러오기
# excel_file = r'C:\Users\PHJ\output\개인정보 운영대장.xlsx'
# wb = load_workbook(excel_file)
# ws = wb['개인정보 추출 및 이용 관리']  # 데이터가 입력될 워크시트 이름

# # 데이터 입력 시작 행 결정 (예: 2행부터 데이터 입력)
# start_row = 2

# # 데이터프레임의 열 순서 조정 (엑셀의 열 순서와 일치하도록)
# df = df[['결재일', '년', '월', '일', '주차', '법인명', '문서번호', '제목', '업무 유형', '추출 위치', '담당 부서', '신청자', '합의 담당자', '링크', '진행 구분']]

# # 데이터프레임을 엑셀 워크시트에 쓰기
# for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=False), start=start_row):
#     for c_idx, value in enumerate(row, start=1):
#         ws.cell(row=r_idx, column=c_idx, value=value)

# # 엑셀 파일 저장
# wb.save(excel_file)

######################################2안 끝##############################################


# 브라우저 종료
driver.quit()
