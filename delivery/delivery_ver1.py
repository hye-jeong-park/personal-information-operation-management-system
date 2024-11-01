from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import time
import traceback
import getpass
from openpyxl import load_workbook
import sys
import re

# 웹드라이버 설정
driver = webdriver.Chrome()

try:
    # 로그인 페이지로 이동
    driver.get('https://gw.com2us.com/')
    
    # 로그인 처리
    username_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'Username'))
    )
    password_input = driver.find_element(By.ID, 'Password')
    
    username = input('아이디를 입력하세요: ')
    password = getpass.getpass('비밀번호를 입력하세요: ')
    
    username_input.send_keys(username)
    password_input.send_keys(password)
    
    login_button = driver.find_element(By.CLASS_NAME, 'btnLogin')
    login_button.click()
    
    # 로그인 성공 여부 확인
    time.sleep(5)
    current_url = driver.current_url
    print(f"로그인 후 현재 URL: {current_url}")
    
    if 'login' in current_url.lower():
        print("로그인에 실패하였습니다.")
        driver.quit()
        sys.exit()
    
    # 업무지원 > 개인정보 파일 전송 페이지로 이동
    driver.get('https://gw.com2us.com/emate_app/00001/bbs/b2307140306.nsf/view?readform&viewname=view01')
    
    # 페이지 이동 후 현재 URL 출력
    print(f"페이지 이동 후 현재 URL: {driver.current_url}")
    
    # 페이지 로딩 대기
    try:
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, 'dhx_skyblue')))
    except Exception as e:
        print("게시글 목록을 찾을 수 없습니다.")
        print(e)
        driver.quit()
        sys.exit()
        
    # 게시글 목록 가져오기
    posts = driver.find_elements(By.XPATH, '//tr[contains(@class, "dhx_skyblue")]')
    print(f"게시글 수: {len(posts)}")
    
    data_list = []
    
    for i in range(len(posts)):
        # 게시글 목록을 다시 가져옵니다. (동적 페이지일 경우 필요)
        posts = driver.find_elements(By.CSS_SELECTOR, 'tr[class*="dhx_skyblue"]')
        post = posts[i]
    
        try:
            # 해당 행의 모든 td 요소를 가져옵니다.
            tds = post.find_elements(By.TAG_NAME, 'td')
    
            # 등록일 추출 (5번째 td)
            등록일_td = tds[4]  # 0-based index
            등록일_text = 등록일_td.get_attribute('title').strip() if 등록일_td.get_attribute('title') else 등록일_td.text.strip()
    
            # 작성자 추출 (3번째 td)
            작성자_td = tds[2]
            작성자 = 작성자_td.find_element(By.TAG_NAME, 'span').text.strip()
    
        except Exception as e:
            print(f"목록에서 데이터 추출 중 오류 발생: {e}")
            등록일_text = 작성자 = ''
            continue  # 오류 발생 시 다음 게시글로 이동
    
        # 요소가 화면에 보이도록 스크롤합니다.
        driver.execute_script("arguments[0].scrollIntoView();", post)
    
        # 클릭 가능할 때까지 대기합니다.
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable(post))
    
        # 게시글 클릭하여 팝업 열기
        post.click()
    
        # 새로운 창으로 전환
        WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
        driver.switch_to.window(driver.window_handles[-1])
    
        # 페이지 로딩 대기
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'AppLineArea')))
    
        try:
            # 상세 페이지에서 제목 확인
            h2_element = driver.find_element(By.CSS_SELECTOR, '#AppLineArea h2')
            h2_text = h2_element.text.strip()
    
            # 제목이 '개인정보 추출 신청서'가 아닌 경우 건너뜀
            if '개인정보 추출 신청서' not in h2_text:
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                continue  # 다음 게시글로 이동
    
            # 현재 창 제목 출력
            print(f"현재 창 제목: {driver.title}")
    
            # 상세 페이지의 테이블 행들 가져오기
            table_rows = driver.find_elements(By.XPATH, '//tbody/tr')
    
            # 초기화
            법인명 = 제목 = 파일형식 = 파일용량 = 개인정보_수 = 진행_구분 = ''
            링크 = driver.current_url
    
            for row in table_rows:
                try:
                    header = row.find_element(By.XPATH, './td[1]').text.strip()
                    value_td = row.find_element(By.XPATH, './td[2]')
                    
                    if '수신자 (부서, 이름)' in header:
                        # 법인명 추출: "컴투스 운영지원, 홍길동" 중 "컴투스"만 추출
                        full_text = value_td.text.strip()
                        법인명 = full_text.split()[0].split(',')[0] if ',' in full_text else full_text.split()[0]
                    
                    elif '제목' in header:
                        # 제목 추출
                        제목 = row.find_element(By.ID, 'DisSubject').text.strip()
                    
                    elif '파밀명 및 용량 (KB)' in header:
                        # 파일형식 및 파일 용량 추출
                        file_info = value_td.text.strip()
                        # 예시: "(Confidential)_20241017_103738_smon_lms_target_list.zip (221KB)"
                        file_match = re.match(r'\(.*?\)_(.*?)\.(zip|xlsx).*?\((\d+KB)\)', file_info)
                        if file_match:
                            filename = file_match.group(1) + '.' + file_match.group(2)
                            파일용량 = file_match.group(3)
                            if filename.endswith('.zip'):
                                파일형식 = 'Zip'
                            elif filename.endswith('.xlsx'):
                                파일형식 = 'Excel'
                            else:
                                파일형식 = ''
                        else:
                            # 다른 형식의 파일명이 있을 경우
                            p_tags = value_td.find_elements(By.TAG_NAME, 'p')
                            if len(p_tags) >= 2:
                                filename = p_tags[0].text.strip()
                                파일용량 = p_tags[1].text.strip()
                                if filename.endswith('.zip'):
                                    파일형식 = 'Zip'
                                elif filename.endswith('.xlsx'):
                                    파일형식 = 'Excel'
                                else:
                                    파일형식 = ''
                            else:
                                파일형식 = ''
                                파일용량 = ''
                    
                    elif '추출된 항목 및 건수' in header:
                        # 개인정보(수) 추출: "한국 휴대전화번호 : 10,773건"에서 "10,773" 추출
                        items = value_td.text.strip()
                        개인정보_수_match = re.search(r'(\d{1,3}(?:,\d{3})*)건', items)
                        개인정보_수 = 개인정보_수_match.group(1) if 개인정보_수_match else ''
                    
                except Exception as e:
                    print(f"상세 페이지에서 특정 데이터 추출 중 오류 발생: {e}")
                    continue
    
            # 진행 구분 설정: '제목'에 '추출완료일' 포함 시 "다운 완료"
            if '추출완료일' in 제목:
                진행_구분 = '다운 완료'
            else:
                진행_구분 = ''
    
            # 데이터 저장
            data = {
                '등록일': 등록일_text,
                '법인명': 법인명,
                '제목': 제목,
                '작성자': 작성자,
                '링크': 링크,
                '파일형식': 파일형식,
                '파일 용량': 파일용량,
                '고유식별정보(수)': '',  # 공백으로 저장
                '개인정보(수)': 개인정보_수,
                '진행 구분': 진행_구분
            }
            data_list.append(data)
        except Exception as e:
            print(f"데이터 추출 중 오류 발생: {e}")
            traceback.print_exc()
    
        # 창 닫기 및 원래 창으로 전환
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
    
        # 잠시 대기
        time.sleep(2)
    
    # 데이터프레임 생성
    df = pd.DataFrame(data_list)
    
    ######################################엑셀화##############################################
    
    # 데이터가 있는 경우에만 엑셀 저장 진행
    if not df.empty:
        # 기존 엑셀 파일 불러오기
        excel_file = r'C:\Users\PHJ\output\개인정보 운영대장.xlsx'
        wb = load_workbook(excel_file)
        ws = wb['개인정보 추출 및 전달']  # 데이터가 입력될 워크시트 이름
    
        # 'S' 열(등록일)에서 데이터가 있는 마지막 행 찾기
        last_row = ws.max_row
        while last_row >= 5:  # 데이터가 시작되는 행 번호는 5
            if ws.cell(row=last_row, column=19).value is not None:  # S열 (등록일) 확인
                break
            last_row -= 1
    
        # 새로운 데이터 입력 시작 행
        if last_row < 5:
            start_row = 5  # 데이터 시작 행
        else:
            start_row = last_row + 1
    
        # 데이터프레임의 열 순서 조정 (엑셀의 열 순서와 일치하도록)
        df = df[['등록일', '법인명', '제목', '작성자', '링크', '파일형식', '파일 용량', '고유식별정보(수)', '개인정보(수)', '진행 구분']]
    
        # 열 매핑 설정 (데이터프레임 열 이름과 엑셀 열 인덱스 매핑)
        column_mapping = {
            '등록일': 19,          # S
            '법인명': 20,          # T
            '제목': 21,            # U
            '작성자': 22,          # V
            '링크': 23,            # W
            '파일형식': 24,        # X
            '파일 용량': 25,       # Y
            '고유식별정보(수)': 26, # Z
            '개인정보(수)': 27,    # AA
            '진행 구분': 28        # AB
        }
    
        # 데이터프레임을 엑셀 워크시트에 쓰기
        for idx, row in df.iterrows():
            # 각 열에 데이터 입력
            for col_name, col_idx in column_mapping.items():
                value = row[col_name]
                ws.cell(row=start_row, column=col_idx, value=value)
            start_row += 1
    
        # 엑셀 파일 저장
        wb.save(excel_file)
    else:
        print("추출된 데이터가 없습니다.")

except Exception as main_e:
    print(f"전체 프로세스 중 오류 발생: {main_e}")
    traceback.print_exc()
finally:
    # 브라우저 종료
    driver.quit()